# Python program to demonstrate
# langdetect

from openpyxl import load_workbook

import sys
sys.stdout.reconfigure(encoding='utf-8')
import regex as re
import pandas as pd

FILE_NAME = "Report_OS_Team.xlsx"
ROOT_PATH = "D:\\_workspace\\project\\Gateway_library\\99_Users\\ToanPham\\r-car-s4-gateway-sw"
FILE_PATH_COLUMN = 1
FILE_NAME_COLUMN = 2
FLAG_COLUMN = 5
EMPTY_CELL = ''

START_ROW=1
END_ROW=8286

PATTERN = re.compile(r'([\p{IsHan}\p{IsBopo}\p{IsHira}\p{IsKatakana}]+)', re.UNICODE)
RE_EXTENSION = r'\w*\.(xlsm|xlsx)'

filename=''
pathfile=''
currentOption=''
currentLine=0
excelFile = pd.ExcelFile(FILE_NAME)
	


dataframe = [];
with open(FILE_NAME, 'rb') as f:
			dataframe.append(pd.read_excel(f,excelFile.sheet_names[1] ,na_filter = False,header=None))


# print(dataframe)
# row column
# print(dataframe[0].iloc[2][1])

for currentRow in range(START_ROW, 4):
	# set variables
	try:
		curentPath = dataframe[0].iloc[currentRow][FILE_PATH_COLUMN]
		curentFile = dataframe[0].iloc[currentRow][FILE_NAME_COLUMN]
		curentFlag = dataframe[0].iloc[currentRow][FLAG_COLUMN]
	except:
		print("CELL out of range")
		continue

	# check file is excel
	if not re.search(RE_EXTENSION, curentFile):
		print("continue")
		continue

	print(curentPath)
	print(curentFile)
	print(curentFlag, EMPTY_CELL)

	#currentfilePath = ROOT_PATH+'\\' +curentPath + '\\' + curentFile;
	currentfilePath = curentFile;
	print(currentfilePath)
	# Do NOT overwrite existing value
	if(curentFlag != EMPTY_CELL):
		print("EMPTY")
		continue

	# read current file
	try:
		currentFileExcel = pd.ExcelFile(currentfilePath)
	except:
		print("file is not exist",currentfilePath)
		
		continue

	try:
     
		with open(currentfilePath, 'rb') as f:
			for sheetName in currentFileExcel.sheet_names:
				print("SHEET", sheetName)
				currentFileDataFrame =  pd.read_excel(f, sheetName)



				# print(currentFileDataFrame)
				# print(currentFileDataFrame.to_string())
				currentFileStr = currentFileDataFrame.to_string()

				# check file if contains japanese
				try:
					if PATTERN.search(currentFileStr):
						print("True JA")

						#load excel file
						workbook = load_workbook(filename=FILE_NAME)
						print(currentfilePath)
						#Pick the sheet "new_sheet"
						print("sheet ",excelFile.sheet_names[1])
						# ws4 = workbook["Task1_Detail"]
						ws4 = workbook[excelFile.sheet_names[1]]
						#modify the desired cell
						ws4.cell(row = (currentRow+1), column = 6).value = "yes"
						#save the file
						workbook.save(filename=FILE_NAME)
						print("SAVED")
						break
					else:
						print("false")
				except Exception as e:
					print(e)
					print("This row throws and error:")
	except:
		print("File is not exist: ",curentFile)

