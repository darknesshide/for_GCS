
# Python program to demonstrate
# langdetect

from openpyxl import load_workbook
from langdetect import detect
import sys
sys.stdout.reconfigure(encoding='utf-8')
import regex as re
import pandas as pd

FILE_NAME = "Report_OS_Team.xlsx"
ROOT_PATH = "D:\\_workspace\\project\\Gateway_library\\99_Users\\ToanPham\\r-car-s4-gateway-sw"
FILE_PATH_COLUMN = 1
FILE_NAME_COLUMN = 2
FLAG_COLUMN = 5
EMPTY_CELL = ''

START_ROW=1
END_ROW=8286

PATTERN = re.compile(r'([\p{IsHan}\p{IsBopo}\p{IsHira}\p{IsKatakana}]+)', re.UNICODE)
RE_EXTENSION = r'\w*\.(js|html)'

filename=''
pathfile=''
currentOption=''
currentLine=0
excelFile = pd.ExcelFile(FILE_NAME)
	


dataframe = [];
with open(FILE_NAME, 'rb') as f:
			dataframe.append(pd.read_excel(f,excelFile.sheet_names[1] ,na_filter = False,header=None))


# print(dataframe)
# row column
# print(dataframe[0].iloc[2][1])

for currentRow in range(START_ROW, 4):
	# set variables
	try:
		curentPath = dataframe[0].iloc[currentRow][FILE_PATH_COLUMN]
		curentFile = dataframe[0].iloc[currentRow][FILE_NAME_COLUMN]
		curentFlag = dataframe[0].iloc[currentRow][FLAG_COLUMN]
	except:
		print("CELL out of range")
		continue

	# check file is excel
	if not re.search(RE_EXTENSION, curentFile):
		print("EXTENSION continue")
		continue

	print(curentPath)
	print(curentFile)
	print(curentFlag, EMPTY_CELL)

	#currentfilePath = ROOT_PATH+'\\' +curentPath + '\\' + curentFile;
	currentfilePath = curentFile;
	print(currentfilePath)
	# Do NOT overwrite existing value
	if(curentFlag != EMPTY_CELL):
		print("EMPTY")
		continue


	try:
		flag_ret= 0
		with open(currentfilePath, 'rb') as f:
			currentFileStr =  f.read().decode('utf-8')
			print(currentFileStr)
			# print(currentFileDataFrame)
			# print(currentFileDataFrame.to_string())
			# currentFileStr = currentFileDataFrame.to_string()

			# check file if contains japanese
			try:
				if PATTERN.search(currentFileStr):
					print("True JA")
					flag_ret = 1
					#load excel file
					workbook = load_workbook(filename=FILE_NAME)
					print(currentfilePath)
					#Pick the sheet "new_sheet"
					print("sheet ",excelFile.sheet_names[1])
					# ws4 = workbook["Task1_Detail"]
					ws4 = workbook[excelFile.sheet_names[1]]
					#modify the desired cell
					# casause of difference of pandas and openpyxl
					ws4.cell(row = (currentRow+1), column = FLAG_COLUMN+1).value = "yes"
					#save the file
					workbook.save(filename=FILE_NAME)
					print("SAVED")
					break
				else:
					print("false")
			except Exception as e:
				print("This row throws and error:")
				print(e)
				continue
		if flag_ret == 0:
			#load excel file
			workbook = load_workbook(filename=FILE_NAME)
			print(currentfilePath)
			#Pick the sheet "new_sheet"
			print("sheet ",excelFile.sheet_names[1])
			# ws4 = workbook["Task1_Detail"]
			ws4 = workbook[excelFile.sheet_names[1]]
			#modify the desired cell
			ws4.cell(row = (currentRow+1), column = FLAG_COLUMN).value = "no"
			#save the file
			workbook.save(filename=FILE_NAME)
			print("SAVED")

	except Exception as e:
		print(e)
		print("File is not exist: ",curentFile)
		continue

